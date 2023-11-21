# 获取股票基金市场交易量情况
import json
import requests
import datetime
import openpyxl
import pandas as pd
import chinese_calendar as cal
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 获取上一个交易日
def get_previous_trading_day(date):
    previous_day = date - datetime.timedelta(days=1)
    while not cal.is_workday(previous_day) or previous_day.weekday() in [5, 6]:
        previous_day -= datetime.timedelta(days=1)
    return previous_day

current_date = datetime.date.today()  # 获取当前日期
previous_day = get_previous_trading_day(current_date)
if cal.is_workday(previous_day) and previous_day.weekday() not in [5, 6]:
    trade_date = previous_day
else:
    trade_date = get_previous_trading_day(previous_day)


""" # 补充某一天的交易量数据
from datetime import datetime
trade_date_str = '2023-11-09'
trade_date = datetime.strptime(trade_date_str, '%Y-%m-%d').date() """


# 获取上交所交易量
Cookie = "ba17301551dcbaf9_gdp_user_key=; gdp_user_id=gioenc-c1785dd3%2Cb695%2C51cg%2C84gg%2C8e7b7b4b9g6b; ba17301551dcbaf9_gdp_session_id_3df712f8-0c77-4c37-bd5f-d89df676be39=true; ba17301551dcbaf9_gdp_session_id_7a21bda2-5089-44d9-add4-c88e9dbb143d=true; ba17301551dcbaf9_gdp_session_id_add47fce-aded-4899-a646-004dc0b42989=true; ba17301551dcbaf9_gdp_session_id_774dd4fd-d40b-4566-bdaf-97ec7d087cfc=true; ba17301551dcbaf9_gdp_session_id_e61de0fe-cdac-4c6e-b5ba-a1e69bc19faf=true; ba17301551dcbaf9_gdp_session_id=692ed08a-e5a6-4cff-93b8-b48059a48974; ba17301551dcbaf9_gdp_session_id_692ed08a-e5a6-4cff-93b8-b48059a48974=true; ba17301551dcbaf9_gdp_sequence_ids={%22globalKey%22:122%2C%22VISIT%22:7%2C%22PAGE%22:35%2C%22VIEW_CLICK%22:70%2C%22CUSTOM%22:12%2C%22VIEW_CHANGE%22:2}"
sse_url = "http://query.sse.com.cn/commonQuery.do?jsonCallBack=jsonpCallback82969584&isPagination=false&sqlId=COMMON_SSE_SJ_SCGM_C&TRADE_DATE="+trade_date.strftime("%Y-%m-%d")
sse_headers = {
    'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
    'Cookie': Cookie,
    'Connection': 'keep-alive',
    'Accept': '*/*',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Host': 'query.sse.com.cn',
    'Referer': 'http://www.sse.com.cn/'
}

sse_data = requests.get(sse_url,headers=sse_headers).text

# 去除多余的回调函数部分，只保留 JSON 数据
sse_parsed_data = json.loads(sse_data[sse_data.index('{'):-1])

sse_gpcjje = None
sse_jjcjje = None

for item in sse_parsed_data['result']:
    if item['PRODUCT_NAME'] == '股票':
        sse_gpcjje = item['TOTAL_TRADE_AMT']
    elif item['PRODUCT_NAME'] == '基金':
        sse_jjcjje = item['TOTAL_TRADE_AMT']


# 获取深交所交易量
szse_url = "https://www.szse.cn/api/report/ShowReport/data?SHOWTYPE=JSON&CATALOGID=1803_sczm&TABKEY=tab1&txtQueryDate="+trade_date.strftime("%Y-%m-%d")
szse_headers = {
    'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
    'Content-Type':'application/json',
    'Connection': 'keep-alive',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate,br',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Host': 'www.szse.cn',
    'Referer': 'https://www.szse.cn/market/overview/index.html',
    'Sec-Ch-Ua': '"Chromium";v="118", "Google Chrome";v="118", "Not=A?Brand";v="99"',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'X-Request-Type':'ajax',
    'X-Requested-With':'XMLHttpRequest'
}
szse_data = requests.get(szse_url,headers=szse_headers).text
szse_data = json.loads(szse_data)
szse_parsed_data = szse_data[0]["data"]

szse_gpcjje = None
szse_jjcjje = None

# 遍历数据
for item in szse_parsed_data:
    if item["lbmc"] == "股票":
        szse_gpcjje = item['cjje'] 
    elif item["lbmc"] == "基金":
        szse_jjcjje = item["cjje"]


# 获取北交所交易量
bse_url = f"https://www.bse.cn/marketStatController/dailyReport.do?callback=jQuery331_169890708570"
bse_headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate,br',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Cache-Control':'max-age=0',
    'Connection': 'keep-alive',
    'Cookie':'Hm_lvt_ef6193a308904a92936b38108b93bd7f=1698907078; Hm_lpvt_ef6193a308904a92936b38108b93bd7f=1698908622',
    'Host': 'www.bse.cn',
    'Sec-Ch-Ua':'"Chromium";v="118", "Google Chrome";v="118", "Not=A?Brand";v="99"',
    'Sec-Ch-Ua-Mobile':'?0',
    'Sec-Ch-Ua-Platform':'"Windows"',
    'Sec-Fetch-Dest':'document',
    'Sec-Fetch-Mode':'navigate',
    'Sec-Fetch-Site':'none',
    'Sec-Fetch-User':'?1',
    'Upgrade-Insecure-Requests':'1',
    'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
}
# 发送HTTP请求
bse_data = requests.get(bse_url,headers=bse_headers).text

# 解析JSON数据
bse_parsed_data = json.loads(bse_data[bse_data.find('['):bse_data.rfind(']')+1])

bse_gpcjje = None
target_date = trade_date.strftime("%Y%m%d")

for entry in bse_parsed_data:
    if entry["rq"] == target_date and entry["xxzrlx"] == '2':
        bse_gpcjje = entry["hqcjje"]

# 将交易量单位统一为万元
sse_gpcjje = round(float(sse_gpcjje.replace(',', '')) * 10000,0)
sse_jjcjje = round(float(sse_jjcjje.replace(',', '')) * 10000,0)
szse_gpcjje = round(float(szse_gpcjje.replace(',', '')) * 10000,0)
szse_jjcjje = round(float(szse_jjcjje.replace(',', '')) * 10000,0)
bse_gpcjje = round(bse_gpcjje/10000,2)

# 将数据写入Excel文件
data = {
    '交易日期': [trade_date],
    '上交所（股票）': [sse_gpcjje],
    '上交所（基金）': [sse_jjcjje],
    '深交所（股票）': [szse_gpcjje],
    '深交所（基金）': [szse_jjcjje],
    '北交所（股票）': [bse_gpcjje],
    '合计':[sse_gpcjje+sse_jjcjje+szse_gpcjje+szse_jjcjje+bse_gpcjje]
}
df = pd.DataFrame(data)


# 读取现有的Excel文件，如果文件不存在，则创建一个新的工作簿
try:
    workbook = openpyxl.load_workbook('市场交易量.xlsx')
except FileNotFoundError:
    workbook = openpyxl.Workbook()

# 选择要写入的工作表，如果不存在，则创建一个新的工作表
sheet_name = 'sheet1'
if sheet_name not in workbook.sheetnames:
    workbook.create_sheet(sheet_name)

# 获取选定的工作表
sheet = workbook[sheet_name]

alignment = Alignment(horizontal='center', vertical='center')
# 将DataFrame数据逐行写入工作表
for row in dataframe_to_rows(df, index=False, header=False):
    sheet.append(row)
   
# 获取最后一行的范围
last_row = sheet.max_row

# 应用居中对齐样式到每个单元格
for row in sheet.iter_rows(min_row=1, max_row=last_row):
    for cell in row:
        cell.alignment = alignment       

# 保存修改后的Excel文件
workbook.save('市场交易量.xlsx')
print(str(trade_date)+'市场交易量爬取成功！')


